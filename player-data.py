import sys
import openpyxl as px
import requests
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup

def teamcrawling(division):
  if division == 'J1':
    id = '#footerj1'
  elif division == 'J2':
    id = '#footerj2'
  elif division == 'J3':
    id = '#footerj3'

  req = requests.get('https://www.football-lab.jp/')
  top_soup = BeautifulSoup(req.text, 'html.parser')
  html_attributes = top_soup.select('{} li a'.format(id))
  team_permalinks = []
  for ha in html_attributes:
    team_permalinks.append(ha.get('href').strip('/'))

  return team_permalinks

def pagecrawling(soup):
    table = soup.select('.statsTbl10 tr')
    standing = []
    for row in table:
        tmp = []
        for item in row.find_all('td'):
            if item.a:
                tmp.append(item.text[0:len(item.text)])
            else:
                tmp.append(item.text)

        standing.append(tmp)

    #配列の要素数をカウント
    length = len(standing)
    #開始位置を指定
    n = 0
    #分割する変数の個数を指定
    s = 16
    #配列を指定した個数で分割していくループ処理
    for i in standing:
        # print(standing[n:n+s:1])
        n += s
        #カウント数が配列の長さを超えたらループ終了
        if n >= length:
            break

    del standing[0:2]

    # チーム名を取得する
    team_name = soup.find('div', id='teamHeader').find('span', class_='jpn').text
    data = { team_name: standing }

    return data

def write_excel(team_data_list):
    # エクセルを取得
    wb = px.Workbook()
    ws = wb.active

    # エクセル2行目以降に取得したデータを出力
    for i, data in enumerate(team_data_list):
      for team_name, player_data in data.items():
        # データを出力するチーム名のシートを作成
        wb.create_sheet(title=team_name)
        ws = wb[team_name]
        # エクセル1行目のヘッダーを出力
        headers = ['ポジション', '背番号', '選手名', '出場', '先発', '出場時間', '攻撃', 'パス', 'クロス', 'ドリブル', 'パスレシーブ', 'シュート', 'ゴール', '奪取', '守備', 'セーブ']
        for i, header in enumerate(headers):
            # ヘッダー値を設定
            ws.cell(row=1, column=1+i, value=header)
            # セルを塗りつぶす
            fill = PatternFill(patternType='solid', fgColor='e0e0e0', bgColor='e0e0e0')
            ws.cell(row=1, column=1+i).fill = fill

        for n, pd in enumerate(player_data):
          for y, d in enumerate(pd):
            ws.cell(row= 2+n, column= 1+y, value= d)

    # 不要なSheetシートを削除
    del wb['Sheet']
    # エクセルファイルの保存
    filename = 'J_player_data.xlsx'
    wb.save(filename)

def main():
  divisions = { 1: 'J1', 2: 'J2', 3: 'J3' }
  print('データを取得したいリーグを選択してください。')
  for key, value in divisions.items():
    print('{0}: {1}'.format(key, value))

  try:
    user_input = int(input())
  except KeyboardInterrupt:
    sys.exit()

  try:
    team_permalinks = teamcrawling(divisions[user_input])
  except:
    print('半角で「1」「2」「3」のいづれかを入力してください。')
    sys.exit()

  team_data_list = []
  for tp in team_permalinks:
    url = 'https://www.football-lab.jp/' + tp;

    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'html.parser')

    # 各チームのデータを「team_data_list」に格納
    team_data_list.append(pagecrawling(soup))

  # エクセルデータを作成
  write_excel(team_data_list)

if __name__ == '__main__':
  main()
