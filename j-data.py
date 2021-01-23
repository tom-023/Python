import sys
import openpyxl as px
import requests
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup

def teamcrawling(division):
  if division == 'J1':
    id = '#footerj1'
  elif division == 'J2':
    id = '#footerj2'
  elif division == 'J3':
    id = '#footerj3'

  req = requests.get('https://www.football-lab.jp/')
  top_soup = BeautifulSoup(req.text, 'html.parser')
  html_attributes = top_soup.select('{} li a'.format(id))
  team_permalinks = []
  for ha in html_attributes:
    team_permalinks.append(ha.get('href').strip('/'))

  return team_permalinks

def pagecrawling(soup):
    data_list = []
    # チーム名を取得する
    team_name = soup.find('div', id='teamHeader').find('span', class_='jpn')
    data_list.append(team_name.text)

    # プログラム2-1｜<h3>タグのなかで、class='boxHeader'のものを変数productsに格納
    products = soup.find_all('div', class_=['statsTbl4', 'statsTbl3'])
    for product in products:
        team_datas = product.find_all('span', class_='numL')
        for td in team_datas:
          data_list.append(td.text)

    return data_list

def write_excel(team_data_list):
    # エクセルを取得
    wb = px.Workbook()
    ws = wb.active

    # エクセルのヘッダーの背景色を設定
    fill = PatternFill(patternType='solid', fgColor='e0e0e0', bgColor='e0e0e0')

    # エクセル1行目のヘッダーを出力
    headers = ['チーム', '攻撃', 'パス', 'クロス', 'ドリブル', 'シュート', 'ゴール', '奪取', '守備', 'セーブ']
    for i, header in enumerate(headers):
        # ヘッダー値を設定
        ws.cell(row=1, column=1+i, value=header)
        # セルを塗りつぶす
        ws.cell(row=1, column=1+i).fill = fill

    # エクセル2行目以降に取得したデータを出力
    for i, deta in enumerate(team_data_list):
      for y, d in enumerate(deta):
        ws.cell(row= 2+i, column= 1+y, value= d)

    # プログラム4-6｜エクセルファイルの保存
    filename = 'J-data.xlsx'
    wb.save(filename)


def main():
  divisions = { 1: 'J1', 2: 'J2', 3: 'J3' }
  print('データを取得したいリーグを選択してください。')
  for key, value in divisions.items():
    print('{0}: {1}'.format(key, value))

  try:
    user_input = int(input())
  except KeyboardInterrupt:
    sys.exit()

  try:
    team_permalinks = teamcrawling(divisions[user_input])
  except:
    print('半角で「1」「2」「3」のいづれかを入力してください。')
    sys.exit()

  team_data_list = []
  for tp in team_permalinks:
    url = 'https://www.football-lab.jp/' + tp;

    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'html.parser')

    # 各チームのデータを「team_data_list」に格納
    team_data_list.append(pagecrawling(soup))

  # 「team_data_list」の中身をチェック
  [print(i, data) for i, data in enumerate(team_data_list)]

  # エクセルデータを作成
  write_excel(team_data_list)

if __name__ == '__main__':
  main()
