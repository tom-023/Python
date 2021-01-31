import sys
import openpyxl as px
import requests
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from IPython import embed

def teamcrawling(division):
    if division == 'J1':
        id = '#footerj1'
    elif division == 'J2':
        id = '#footerj2'
    elif division == 'J3':
        id = '#footerj3'

    req = requests.get('https://www.football-lab.jp/')
    top_soup = BeautifulSoup(req.text, 'html.parser')
    html_attributes = top_soup.select('{} li a'.format(id))
    team_permalinks = []
    for ha in html_attributes:
        team_permalinks.append(ha.get('href').strip('/'))

    return team_permalinks


def write_excel(data_list):
    # エクセルを取得
    wb = px.Workbook()
    ws = wb.active

    # エクセルのヘッダーの背景色を設定
    fill = PatternFill(patternType='solid', fgColor='e0e0e0', bgColor='e0e0e0')

    # エクセル1行目のヘッダーを出力
    headers = ['攻撃', 'パス', 'クロス', 'ドリブル', 'シュート', 'ゴール', '奪取', '守備', 'セーブ', 'ゴール期待値', 'シュート', '枠内シュート', 'PKによるシュート', 'パス', 'クロス', '直接ＦＫ', '間接ＦＫ', 'ＣＫ', 'スローイン', 'ドリブル', 'タックル', 'クリア', 'インターセプト', 'オフサイド', '警告', '退場', '３０ｍライン進入', 'ペナルティエリア進入', '攻撃回数', 'チャンス構築率', 'ボール支配率',
               '敵攻撃', '敵パス', '敵クロス', '敵ドリブル', '敵シュート', '敵ゴール', '敵奪取', '敵守備', '敵セーブ', '敵ゴール期待値', '敵シュート', '敵枠内シュート', '敵PKによるシュート', '敵パス', '敵クロス', '敵直接ＦＫ', '敵間接ＦＫ', '敵ＣＫ', '敵スローイン', '敵ドリブル', '敵タックル', '敵クリア', '敵インターセプト', '敵オフサイド', '敵警告', '敵退場', '敵３０ｍライン進入', '敵ペナルティエリア進入', '敵攻撃回数', '敵チャンス構築率', '敵ボール支配率',
               '得点', '失点']
    for i, header in enumerate(headers):
        # ヘッダー値を設定
        ws.cell(row=1, column=1+i, value=header)
        # セルを塗りつぶす
        ws.cell(row=1, column=1+i).fill = fill

    columns_data = []
    print("これからデータを選別します")
    for data in data_list:
        for d in data:
            home_team_cp = d[3:75:8]
            away_team_cp = d[5:77:8]

            home_team_st = d[75:244:8]
            away_team_st = d[77:246:8]

            home_team_goal = [d[-3]]
            away_team_goal = [d[-1]]
            columns_data.append(home_team_cp + home_team_st + away_team_cp + away_team_st + home_team_goal + away_team_goal)

    # %を削除し、データを文字列から数値に変換する
    last_data = [[float(d.replace('%', '')) for d in cd] for cd in columns_data]

    # エクセル2行目以降に取得したデータを出力
    print("これからエクセルに書き出します")
    for i, data in enumerate(last_data):
        for y, d in enumerate(data):
            ws.cell(row= 2+i, column= 1+y, value= d)

    # プログラム4-6｜エクセルファイルの保存
    filename = 'J-battledata.xlsx'
    wb.save(filename)


def main():
    divisions = { 1: 'J1', 2: 'J2', 3: 'J3' }
    print('データを取得したいリーグを選択してください。')
    for key, value in divisions.items():
        print('{0}: {1}'.format(key, value))

    try:
        user_input = int(input())
    except KeyboardInterrupt:
        sys.exit()

    try:
        team_permalinks = teamcrawling(divisions[user_input])
    except:
        print('半角で「1」「2」「3」のいづれかを入力してください。')
        sys.exit()

    data_list = []
    for tp in team_permalinks:
        print('これから{}の処理を行います。'.format(tp))
        url = 'https://www.football-lab.jp/' + tp + '/match';

        r = requests.get(url)
        soup = BeautifulSoup(r.text, 'html.parser')

        html_attributes = soup.select(".statsTbl10 a")
        battle_links_list = []
        for ha in html_attributes:
            battle_links_list.append(ha.get('href'))

        battle_links_list = battle_links_list[1::2]
        team_data_list = []
        for battle_link in battle_links_list:
            link = 'https://www.football-lab.jp' + battle_link
            re = requests.get(link)
            battle_soup = BeautifulSoup(re.text, 'html.parser')
            battle_html_attributes = battle_soup.select(".statsTbl6 td")
            goal_html_attributes = battle_soup.select(".vsHeader .numL")
            battle_html_attributes.extend(goal_html_attributes)
            battle_data_list = []
            for bha in battle_html_attributes:
                battle_data_list.append(bha.text)
            team_data_list.append(battle_data_list)
        data_list.append(team_data_list)

    print("これからエクセル作成処理を行います")
    # エクセルデータを作成
    write_excel(data_list)

if __name__ == '__main__':
    main()
